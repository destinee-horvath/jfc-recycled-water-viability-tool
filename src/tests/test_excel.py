"""
tests/test_excel.py
=====================
PHASE-agnostic: backend.build_xlsx() — colour-coded Excel export.
Verifies valid workbook bytes and that phase/check rows are coloured by
their own state (not just the phase's rolled-up state).
"""

import io

import openpyxl
import pytest

import backend as B
import config as C
from backend.excel import _hex, _tint

VALID_PHASE1 = {
    "water_source": "Treated effluent",
    "supplier_authority": "Council-run scheme",
    "council_approval_held": "Yes",
    "user_approval_state": "Permitted for this use",
}


def _inputs(water_quality_overrides=None, supplier_approval_overrides=None):
    inputs = {
        "supplier_approval": dict(VALID_PHASE1, **(supplier_approval_overrides or {})),
        "water_quality": dict({
            "application_type": "Earthworks (Compaction)",
            "ew_sugar": 50, "ew_oil": 20, "ew_ph": 6, "ew_tds": 1000,
            "ew_chloride": 100, "ew_sulphate": 100, "ew_alkali": 500, "ew_tss": 5000,
        }, **(water_quality_overrides or {})),
        "rwmp": {}, "site_runoff": {}, "soil_conditions": {}, "whs": {}, "financial": {},
    }
    return inputs


def _build(inputs, name="Test Assessment"):
    results = B.run_assessment(inputs)
    readiness = B.readiness_summary(results)
    meta = {"assessment_name": name, "assessed_by": "tester"}
    xlsx_bytes = B.build_xlsx(meta, results, readiness)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    return wb.active, results, readiness


def test_build_xlsx_returns_valid_workbook_bytes():
    ws, results, readiness = _build(_inputs())
    assert ws is not None
    assert ws["A1"].value == C.APP_TITLE


def test_no_final_verdict_row():
    """By design: no headline verdict/score, just blocking/pending/confirmed."""
    ws, results, readiness = _build(_inputs())
    values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Final verdict" not in values


def test_readiness_section_headers_present_and_coloured():
    ws, results, readiness = _build(_inputs())
    values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    row = next(r for r in range(1, ws.max_row + 1)
               if ws.cell(row=r, column=1).value == "Confirmed")
    header = ws.cell(row=row, column=1)
    assert header.fill.fgColor.rgb.endswith(_hex(C.COLOUR_SUCCESS))
    assert "Required before works (blocking)" in values
    assert "Pending actions (conditional)" in values


def test_phase_banner_row_coloured_by_phase_rolled_up_state():
    # ew_sulphate=999 exceeds the Table 3-1 limit -> forces REJECT.
    ws, results, readiness = _build(_inputs(water_quality_overrides={"ew_sulphate": 999}))
    assert results["water_quality"].state == "REJECT"

    banner_row = next(
        r for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(row=r, column=1).value, str)
        and ws.cell(row=r, column=1).value.startswith(C.PHASE_DISPLAY_NAMES["water_quality"])
    )
    banner = ws.cell(row=banner_row, column=1)
    assert "REJECT" in banner.value
    expected = _hex(C.STATE_COLOURS["REJECT"])
    assert banner.fill.fgColor.rgb.endswith(expected)


def test_individual_check_row_coloured_by_its_own_state_not_phase_state():
    """A REJECT check must use the REJECT tint and a PROCEED check the
    PROCEED tint, even within the same REJECT-rolled-up phase."""
    ws, results, readiness = _build(_inputs(water_quality_overrides={"ew_sulphate": 999}))
    assert results["water_quality"].state == "REJECT"

    sulphate_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=2).value == "Sulphate as SO3"
    )
    sugar_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=2).value == "Sugar"
    )
    assert ws.cell(row=sulphate_row, column=3).value == "REJECT"
    assert ws.cell(row=sugar_row, column=3).value == "PROCEED"

    reject_fill = ws.cell(row=sulphate_row, column=1).fill.fgColor.rgb
    proceed_fill = ws.cell(row=sugar_row, column=1).fill.fgColor.rgb
    assert reject_fill != proceed_fill
    assert reject_fill.endswith(_tint(C.STATE_COLOURS["REJECT"]))
    assert proceed_fill.endswith(_tint(C.STATE_COLOURS["PROCEED"]))


def test_handles_phase_with_no_checks_gracefully():
    # build_xlsx must not raise regardless of phase state.
    ws, results, readiness = _build(_inputs())
    assert ws.max_row > 10


def test_tint_blends_toward_white():
    base = "#E24B4A"
    tinted = _tint(base)
    orig = tuple(int(base.lstrip("#")[i:i+2], 16) for i in (0, 2, 4))
    new = tuple(int(tinted[i:i+2], 16) for i in (0, 2, 4))
    assert all(n >= o for n, o in zip(new, orig))


def test_hex_strips_hash_and_uppercases():
    assert _hex("#1d9e75") == "1D9E75"
    assert _hex("1D9E75") == "1D9E75"
