"""
backend/excel.py
==================
COLOUR-CODED EXCEL (.xlsx) EXPORT — a human-readable companion to the plain
CSV export (persistence.py), for opening/printing/sharing rather than
re-importing. The CSV remains the machine-readable, re-importable format
used by "Load selected" / "Load uploaded"; this file is display-only.

Two sheets: "Readiness Summary" (colour-coded per-check states) and, when
raw inputs are supplied, "Assessment Inputs" (Phase | Field | Value — a
readable mirror of the CSV's own phase/field/value columns).
"""

from __future__ import annotations

import io
from datetime import datetime

import config as C
from .models import PhaseResult

_HEADER_FONT_COLOUR = "FFFFFF"
_TEXT_COLOUR = "222222"


def _hex(colour: str) -> str:
    """Strip a leading '#' — openpyxl wants 'RRGGBB', config.py stores '#RRGGBB'."""
    return colour.lstrip("#").upper()


def _tint(colour: str, amount: float = 0.78) -> str:
    """Blend a hex colour toward white for a pastel row fill (readable text,
    still colour-coded). ``amount`` is the fraction of white blended in."""
    h = _hex(colour)
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = round(r + (255 - r) * amount)
    g = round(g + (255 - g) * amount)
    b = round(b + (255 - b) * amount)
    return f"{r:02X}{g:02X}{b:02X}"


def build_xlsx(meta: dict, results: dict[str, PhaseResult], readiness: dict,
                inputs: dict[str, dict] | None = None) -> bytes:
    """
    Build a colour-coded Readiness Summary workbook: one banner row per
    phase (filled with that phase's PROCEED/CONDITIONAL/REJECT/NA colour)
    followed by one row per check (filled with that check's own state
    colour, so a REJECT check stands out even inside an otherwise
    CONDITIONAL phase). No headline verdict/score by design — just the
    blocking/pending/confirmed breakdown.

    If ``inputs`` (the raw {phase: {field: value}} dict) is supplied, a
    second "Assessment Inputs" sheet is added — see module docstring.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Readiness Summary"

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    col_widths = {"A": 30, "B": 44, "C": 14, "D": 38, "E": 52}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    row = 1

    # --- Title -------------------------------------------------------------
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = C.APP_TITLE
    cell.font = Font(bold=True, size=14, color=_hex(C.COLOUR_PRIMARY))
    row += 2

    # --- Assessment metadata -------------------------------------------------
    meta_rows = [
        ("Assessment name", meta.get("assessment_name", "")),
        ("Assessed by", meta.get("assessed_by", "")),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in meta_rows:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, color=_TEXT_COLOUR)
        ws[f"B{row}"] = value
        row += 1
    row += 1

    # --- Column headers ------------------------------------------------------
    headers = ["Phase", "Check", "State", "Reference", "Detail"]
    for col_idx, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=text)
        c.font = Font(bold=True, color=_HEADER_FONT_COLOUR)
        c.fill = PatternFill("solid", fgColor=_hex(C.COLOUR_PRIMARY))
        c.border = border
    ws.freeze_panes = f"A{row + 1}"
    row += 1

    # --- One banner row per phase, one data row per check -------------------
    for phase in C.PHASES:
        r = results.get(phase["id"])
        if not r:
            continue

        phase_colour = C.STATE_COLOURS.get(r.state, C.COLOUR_NEUTRAL)
        ws.merge_cells(f"A{row}:E{row}")
        banner = ws[f"A{row}"]
        banner.value = f"{phase['label']}  —  {r.state}"
        banner.font = Font(bold=True, color=_HEADER_FONT_COLOUR)
        banner.fill = PatternFill("solid", fgColor=_hex(phase_colour))
        banner.border = border
        row += 1

        if not r.checks:
            ws.cell(row=row, column=1, value="(no checks recorded)").font = \
                Font(italic=True, color=_hex(C.COLOUR_NEUTRAL))
            row += 1
            continue

        for check in r.checks:
            check_colour = C.STATE_COLOURS.get(check.state, C.COLOUR_NEUTRAL)
            fill = PatternFill("solid", fgColor=_tint(check_colour))
            values = ["", check.label, check.state, check.reference, check.detail]
            for col_idx, value in enumerate(values, start=1):
                c = ws.cell(row=row, column=col_idx, value=value)
                c.fill = fill
                c.border = border
                c.alignment = wrap
                if col_idx == 3:
                    c.font = Font(bold=True, color=_TEXT_COLOUR)
            row += 1
        row += 1  # blank spacer row between phases

    # --- Readiness lists (blocking / pending / confirmed) --------------------
    sections = [
        ("Required before works (blocking)", readiness.get("blocking") or [], C.COLOUR_DANGER),
        ("Pending actions (conditional)", readiness.get("pending") or [], C.COLOUR_WARNING),
        ("Confirmed", readiness.get("confirmed") or [], C.COLOUR_SUCCESS),
    ]
    for title, checks, colour in sections:
        ws.merge_cells(f"A{row}:E{row}")
        header = ws[f"A{row}"]
        header.value = title
        header.font = Font(bold=True, color=_HEADER_FONT_COLOUR)
        header.fill = PatternFill("solid", fgColor=_hex(colour))
        row += 1
        if not checks:
            ws.cell(row=row, column=1, value="(none)").font = Font(italic=True)
            row += 1
        else:
            for check in checks:
                ws.cell(row=row, column=1, value=check.label).fill = \
                    PatternFill("solid", fgColor=_tint(colour))
                row += 1
        row += 1

    # --- Disclaimer -----------------------------------------------------------
    ws.merge_cells(f"A{row}:E{row}")
    disclaimer = ws[f"A{row}"]
    disclaimer.value = C.DISCLAIMER
    disclaimer.font = Font(italic=True, size=8, color=_hex(C.COLOUR_NEUTRAL))
    disclaimer.alignment = wrap
    ws.row_dimensions[row].height = 40
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    buffer_note = ws[f"A{row}"]
    buffer_note.value = C.AGWR_BUFFER_ZONE_LIMITATION_NOTE
    buffer_note.font = Font(italic=True, size=8, color=_hex(C.COLOUR_NEUTRAL))
    buffer_note.alignment = wrap
    ws.row_dimensions[row].height = 40

    if inputs:
        _add_inputs_sheet(wb, inputs)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_inputs_sheet(wb, inputs: dict[str, dict]) -> None:
    """Readable Phase | Field | Value dump of every raw input — display-only
    companion to the CSV's own phase/field/value columns (persistence.py)."""
    from openpyxl.styles import Font, PatternFill, Border, Side

    ws = wb.create_sheet("Assessment Inputs")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, width in {"A": 22, "B": 34, "C": 44}.items():
        ws.column_dimensions[col].width = width

    for col_idx, text in enumerate(["Phase", "Field", "Value"], start=1):
        c = ws.cell(row=1, column=col_idx, value=text)
        c.font = Font(bold=True, color=_HEADER_FONT_COLOUR)
        c.fill = PatternFill("solid", fgColor=_hex(C.COLOUR_PRIMARY))
        c.border = border
    ws.freeze_panes = "A2"

    row = 2
    for phase_id, vals in inputs.items():
        for field, value in (vals or {}).items():
            if isinstance(value, (list, tuple)):
                value = ", ".join(str(x) for x in value)
            for col_idx, val in enumerate([phase_id, field, "" if value is None else value], start=1):
                ws.cell(row=row, column=col_idx, value=val).border = border
            row += 1
